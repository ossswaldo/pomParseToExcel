from lxml import etree # pip install lxml==4.3.2
from xml.dom import minidom
from collections import defaultdict
import openpyxl #pip install openpyxl==2.5.14
import xlsxwriter #pip install XlsxWriter==1.2.8
from datetime import datetime
import os
import glob
import time

class pomToExcel():


    def __init__(self):
        self.pom = None
        self.projectName = None
        self.groupId = None
        self.artifactId = None
        self.counter = 0
        self.infoValue = None
        self.last_col_a_value = None
        self.excelCounter = None
        self.xmlCounter = None
        self.elementCounter = 0
        self.chosen_element = None
        self.parseCounter = 0 
        self.repoName = None



    def execute(self):
        start_time = time.time()
        self.fileCounter()
        self.create()
        for k in range(self.xmlCounter):
            self.fileHandler()
            self.parse()
            print()
            k += 1
        print("--- %s seconds to complete ---" % round(time.time() - start_time, 2))


    def parse(self):
        # File that is being read
        pom= self.chosen_element
        # counter is used to determined how many dependencies were found in pom file.
        counter = 0
        # project is used to parse the file to exclusivly find the first artifact ID to find project name
        project = minidom.parse(pom)
        #tree is used to parse the pom files and to iterate trough the file to find the depencies
        #root gets the root of the file
        tree = etree.parse(pom)
        root = tree.getroot()

        #artifactId is used to find the first tag that matches artifactId in the pom files
        #projectName is used to get the data from artifactId
        artifactId = project.getElementsByTagName('artifactId')[0]
        self.projectName = artifactId.firstChild.data

        #depend sets to xpath to fine the depencies
        #dependencyInfo sets an empty dictionary
        depend = root.xpath("//*[local-name()='dependency']")
        dependencyInfo = defaultdict(dict)

        #For loops iterated through the file to store data inside the dictionary
        #data beind retrived is groupId, artifactId and version
        for dep in depend:
            infoList = []
            self.counter += 1
            for child in dep.getchildren():
                infoList.append(child.tag.split('}')[1])
                infoList.append(child.text)

            #list where data is being stored
            dependencyInfo[infoList[1]].update({infoList[2] : infoList[3],infoList[4] : infoList[5]})
                            
        #print statement of all the data
        print(datetime.now(),"""%i Dependency where found in %s's pom file.""" % (self.counter,self.projectName))
        #print(dependencyInfo)
        
        for dependencyId, info in dependencyInfo.items():
            self.parseCounter += 1
            additionalInfo = {}

            for infoName, infoValue in info.items():
                if infoName == "artifactId":
                    self.artifactId = info["artifactId"]
                    self.groupId = dependencyId
                elif infoName == "groupId":
                    self.artifactId = dependencyId
                    self.groupId = info["groupId"]
                else:
                    additionalInfo[infoName] = infoValue
            

            # if self.groupId:
            #     print()
            #     print(f"groupId = {self.groupId}")
            
            # if self.artifactId:
            #     print(f"artifactId = {self.artifactId}")
            
            # for infoName, infoValue in additionalInfo.items():
            #     print(f"{infoName} = {infoValue}")

            self.infoValue = infoValue
            self.excelWriting()


        print(datetime.now(),"%i dependencies where parsed " %self.parseCounter)
        print(datetime.now(),"%i dependencies where written in excel " %self.excelCounter)

    def excelWriting(self):

        self.lastcell()

        #Code gave us a warning error
        # xfile = openpyxl.load_workbook('Libraries.xlsx')
        # sheet = xfile.get_sheet_by_name('Sheet1')
        xfile = openpyxl.load_workbook('Libraries.xlsx')
        xfile.sheetnames
        sheet = xfile["Sheet1"]

        self.excelCounter = 0
        for k in range(0, self.parseCounter):
            i = self.last_col_a_value
            self.excelCounter += 1

            column_cell_reponame= "A"
            column_cell_projectname= "B"
            column_cell_groupID= "C"
            column_cell_artifactId= "D"
            column_cell_Version= "E"

            # sheet[column_cell_reponame+str(i+1)] = self.projectName
            sheet[column_cell_projectname+str(i+1)] = self.projectName
            sheet[column_cell_groupID+str(i+1)] = self.groupId
            sheet[column_cell_artifactId+str(i+1)] = self.artifactId
            sheet[column_cell_Version+str(i+1)] = self.infoValue

        xfile.save('Libraries.xlsx')


    def lastcell(self):

        xfile = openpyxl.load_workbook('Libraries.xlsx')
        xfile.sheetnames
        sheet = xfile["Sheet1"]

        last_row = sheet.max_row
        while sheet.cell(column=2, row=last_row).value is None and last_row > 0:
            last_row -= 1
        self.last_col_a_value = sheet.cell(column=2, row=last_row).row

        xfile.save('Libraries.xlsx')

    def create(self):

        PATH = './Libraries.xlsx'

        print(datetime.now(),"Searching for Libraries.xlsx to write into")
        if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
            print(datetime.now(),"File exists and is readable")
            print(datetime.now(),"Opening File...")
            print()
        else:
            print(datetime.now(),"Either the file is missing or not readable")
            try:
                print(datetime.now(),"Creating excel file.. Libraries.xlsx")
                print()
                # Workbook() takes one, non-optional, argument
                # which is the filename that we want to create.
                workbook = xlsxwriter.Workbook('Libraries.xlsx')

                # The workbook object is then used to add new
                # worksheet via the add_worksheet() method.
                worksheet = workbook.add_worksheet()

                # Use the worksheet object to write
                # data via the write() method.
                worksheet.write('A1', 'Repositoty Name')
                worksheet.write('B1', 'Project Name')
                worksheet.write('C1', 'groupId')
                worksheet.write('D1', 'artifactId')
                worksheet.write('E1', 'Version')

                # Finally, close the Excel file
                # via the close() method.
                workbook.close()
            except:
                pass

    def fileCounter(self):

        ext = "*.xml"
        directory = './'

        self.xmlCounter = len(glob.glob1(directory,ext))
        print(datetime.now(),"%i .xml files where found in directory" %self.xmlCounter)


    def fileHandler(self):

        ext = "*.xml"
        directory = './'

        """ Notes """

        #lists all of the file out
        # for file in glob.glob1(directory,ext):
        #   print(file)

        #  prints out a set with the directory of where its located.
        sets_w_dirPlusName = set(glob.glob(directory+ext)) # Using a comma with give you an error, you need to use a `+`

        #only prints out a set with the files inside.
        sets_w_onlyfile = set(glob.glob(ext))

        #makes a list from the set then sorts it
        my_list = list(sets_w_onlyfile)
        my_list.sort()

        #makes a list from thr set then sorts it
        my_list_w_fileLocation = list(sets_w_dirPlusName)
        my_list_w_fileLocation.sort()



        self.chosen_element = my_list[self.elementCounter]
        PATH = my_list_w_fileLocation[self.elementCounter]

        if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
            print(datetime.now(),""""%s" is loaded and readable""" %self.chosen_element)
            print(datetime.now(),"Opening File...")
        else:
            print(datetime.now(),"Either the file is missing or not readable..trying next file")
            self.fileHandler()


        self.elementCounter += 1


if __name__== "__main__":
    taskMaster =  pomToExcel()
    taskMaster.execute()
